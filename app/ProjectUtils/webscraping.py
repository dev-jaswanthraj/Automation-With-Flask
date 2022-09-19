from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from random import randint
from app import app
from openpyxl import Workbook
from flask_login import current_user


class Scrap:

    def __init__(self, productname):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.filename = str(current_user.id)+str(randint(1, 1000))+str(productname)+".xlsx"
        self.productname = productname
        option = Options()
        option.add_experimental_option("detach", True)
        self.driver = webdriver.Chrome(service=ChromService(ChromeDriverManager().install()), options=option) # Driver


    def close_diver(self):
        self.driver.close()

    def check(self, item, path):
        try:
            __test_data = item.find_element(By.XPATH, path)
            return True
        except:
            return False

    def open_browser_with_amazon(self):
        
        
        self.driver.get("https://www.amazon.in/")

        elem = self.driver.find_element(by=By.NAME, value="field-keywords")
        elem.send_keys(self.productname)
        elem.send_keys(Keys.ENTER)

        items = self.driver.find_elements(By.XPATH, '//div[@data-component-type= "s-search-result"]')

    
        self.ws['A1'] = "Product Name"
        self.ws['B1'] = "Product Price"
        #ws['C1'] = "Discount %"
        self.ws['C1'] = "Image Link"
        for item in items:
            if(self.check(item=item, path = './/span[@class = "a-size-base-plus a-color-base a-text-normal"]')):
                product_name = item.find_element(By.XPATH, './/span[@class = "a-size-base-plus a-color-base a-text-normal"]').text
            elif(self.check(item=item, path='.//span[@class = "a-size-medium a-color-base a-text-normal"]')):
                product_name = item.find_element(By.XPATH, './/span[@class = "a-size-medium a-color-base a-text-normal"]').text
            if(self.check(item=item, path='.//span[@class = "a-price"]')):
                product_price = item.find_element(By.XPATH, './/span[@class = "a-price"]').text
            else:
                product_price = "None"
                #discount = item.find_element(By.TAG_NAME, 'span').text
            if(self.check(item=item, path=".//img[@class = 's-image']")):
                image_link = item.find_element(By.XPATH, ".//img[@class = 's-image']").get_attribute('src')
            product_info = (product_name, product_price, image_link)
            self.ws.append(product_info)
    

        try:
            __path = app.config["UPLOAD_FOLDER"]+"/"+self.filename
            self.wb.save(__path)
            return self.filename
        except:
            return False

    def open_browser_with_chroma(self):

        self.driver.get("https://www.croma.com/")
        elem = self.driver.find_element(by=By.ID, value="search")
        elem.send_keys(self.productname)
        elem.send_keys(Keys.ENTER)

        items = self.driver.find_elements(By.XPATH, '//li[@class= "product-item"]')
        print(items)


# obj = Scrap("Laptop HP")
# obj.open_browser_with_chroma()
#obj.close_diver()
