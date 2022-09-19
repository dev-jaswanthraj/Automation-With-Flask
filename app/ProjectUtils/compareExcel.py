import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from flask_login import current_user
from app import app
from random import randint


class comparefile():

    __alp_dict = {
        1:'A',
        2:'B',
        3:'C',
        4:'D',
        5:'E'
    }

    def __init__(self, file1, file2):
        self.file_name = str(current_user.id)+str(randint(1, 1000))+"CompareResult.xlsx"
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        self.dfvar1 = pd.read_excel(file1)
        self.dfvar1.sort_values(by="Transaction ID", inplace=True)
        self.dfvar2 = pd.read_excel(file2)
        self.dfvar2.sort_values(by="Transaction ID", inplace=True)
        self.ws.append(list(self.dfvar1.columns))
        self.diff_patten = PatternFill('solid', fgColor="f7797d")

    def compare(self):
        __c_row = 2
        diff_cells = []
        for col in range(len(self.dfvar1)):

            __c_col = 1
            d1, d2 = self.dfvar1.iloc[col], self.dfvar2.iloc[col]
            r = []

            for x, y in zip(d1, d2):

                if x != y:
                    r.append("->".join((str(x), str(y))))
                    diff_cells.append(self.__alp_dict[__c_col]+str(__c_row))
                else:
                    r.append(x)

                if __c_col >= 5:
                    __c_col = 1
                else:
                    __c_col += 1

            self.ws.append(r)
            

            __c_row += 1

        for cell in diff_cells:
            c = self.ws[cell]
            c.fill = self.diff_patten 

        
        try:
            __path = app.config['UPLOAD_FOLDER']+"/"+self.file_name
            
            self.wb.save(__path)
            
            return self.file_name
        except:
            return False   
